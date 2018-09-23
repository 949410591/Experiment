from openpyxl import load_workbook
from openpyxl import Workbook

def combine():
	wb = load_workbook('courses.xlsx')
	ws = wb['students']
	wt = wb['time']
	wc = wb.create_sheet("combine")
	for a in list(ws.values)[:]:
		for b in list(wt.values)[:]:
			if a[1] == b[1]:
				a = a + b[-1:]
				wc.append(a)
	wb.save("courses.xlsx")

def split():
	wb = load_workbook('courses.xlsx')
	wc = wb['combine']
	data = {}

	for i in list(wc.values)[1:]:
		year = i[0].year
		if data.setdefault(year,[]):
			data[year].append(i)
		else:
			data[year].append(i)

	for i in data.keys():
		wb1 = Workbook()
		ws = wb1.active
		ws.append(list(wc.values)[0])
		for row in data[i]:
			ws.append(row)
		wb1.save("%d.xlsx"%i)

	
if __name__ == "__main__":
	combine()
	split()


    
